from openpyxl import Workbook

while True:
    wb = Workbook()
    ws = wb.active
    ws.title = "MHW-RTG Power Decrease"

    de = (1 - (1 / 2) ** (1 / 87.7)) * 100  # Ratio of decline per year
    launch_date = 1977
    potencia_inicial = 470
    potencia_final = 0
    row = 3
    cell1 = ""
    cell2 = ""

    ws["A1"] = launch_date
    ws["B1"] = potencia_inicial

    data_calcul = int(input("In which year do you want to run the calculation? Note: To make useful charts in Excel, input distant years, like 2500. "))
    anys_diferencia = data_calcul - launch_date
    filera_maxima = anys_diferencia + 1

    precisio = int(input("How many decimals of precision do you need? "))
    excel = input("Save results in Excel? No --> 0 Yes --> 1: ")
    print("")

    potencia_any_anterior = potencia_inicial * ((100 - de) / 100)
    ws["A2"] = data_calcul - anys_diferencia + 1
    ws["B2"] = round(potencia_any_anterior, precisio)
    anys_diferencia -= 1
    # Block of lines --> To calculate 1st year

    while anys_diferencia > 0:
        potencia_any_anterior = potencia_any_anterior * ((100 - de) / 100)
        cell1 = "A" + str(row)
        cell2 = "B" + str(row)
        ws[cell1] = data_calcul - anys_diferencia + 1
        ws[cell2] = round(potencia_any_anterior, precisio)
        row += 1
        anys_diferencia -= 1
    potencia_final = potencia_any_anterior
    # Block of lines --> To calculate remaining years

    print(f'The initial power at {launch_date} was {potencia_inicial}W')

    potencia_final = round(potencia_final, precisio)
    print(f"Power at {data_calcul}: {potencia_final}W")

    potencia_restant = 100 * potencia_final / potencia_inicial
    potencia_restant = round(potencia_restant, precisio)
    print(f"Percentage of remaining power at {data_calcul} is: {potencia_restant}% \n")

    if excel == "1":
        wb.save("MHT-RTG Power Decrease.xlsx")
